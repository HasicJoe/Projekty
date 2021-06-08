import os
import sys
import re
import pandas
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from datetime import datetime
from modules.values import Values
from modules.frames import Frame
from modules.depo import Depo
from modules.empl import Empl

class Parser():
    def __init__(self, filename,window):
        self.bar_value = 0
        self.rv = 0
        self.window = window
        self.parse_file(filename)
    
    def inc_bar(self):
        self.window.bar['value'] = self.window.bar['value'] + 10
        
    def parse_file(self,filename):
        self.inc_bar()
        
        if not re.match("^.+(xlsx|xls)$", filename.name):
            print(Values.INVALID_FILE)
            self.rv = Values.INVALID_FILE
            return 
           
        spec_columns = ['Meno vodiča', 'Dátum výkonu', 'počet doručených stopov',
        'hmotnosť doručených objednávok', 'počet zvezených stopov', 
        'hmotnosť zvezených objednávok', 'počet stopov rozvoz', 'počet stopov zvoz']
        
        excel_data_df = pandas.read_excel(open(filename.name,'rb'), usecols=spec_columns,
        dtype = {'Meno vodiča': str, 'počet doručených stopov': int,
        'hmotnosť doručených objednávok': float, 'počet zvezených stopov': int,
        'hmotnosť zvezených objednávok': float, 'počet stopov rozvoz': int,
        'počet stopov zvoz': float})
        
        lists = Frame()
        
        for record in excel_data_df.index:
            # update invalid floats
            if pandas.isna(excel_data_df['hmotnosť doručených objednávok'][record]):
                excel_data_df.loc[record,'hmotnosť doručených objednávok'] = 0.0
            if pandas.isna(excel_data_df['hmotnosť zvezených objednávok'][record]):
                excel_data_df.loc[record,'hmotnosť zvezených objednávok'] = 0.0
            
            #append drivers name to list
            if not lists.in_names(excel_data_df['Meno vodiča'][record]):
                lists.add_name(excel_data_df['Meno vodiča'][record])
            
            #update date format - split_date[0] - day, split_date[1] - month, split_date[2] - year
            split_date = re.split('\.', excel_data_df['Dátum výkonu'][record])
            new_format = split_date[2] + '-' + split_date[1] + '-' + split_date[0]
            
            if not lists.in_days(new_format):
                lists.add_day(new_format)
            if not lists.in_years(split_date[2]):
                lists.add_year(split_date[2])
            month_format = split_date[2] + '-' + split_date[1]
            if not lists.in_months(month_format):
                lists.add_month(month_format)
            
            #update date format in 'Dátum výkonu' column
            excel_data_df.loc[record,'Dátum výkonu'] = new_format
                
        self.inc_bar()
        self.parse_data(excel_data_df, lists)
        
        
    def parse_data(self, df, lists):
        depo = Depo()
        depo.parse_depo_daily(df, lists) 
        self.inc_bar()
        depo.parse_depo_monthly(df, lists)
        self.inc_bar()
        empl = Empl()
        empl.parse_empl(df, lists)
        self.inc_bar()
        self.create_sheet(depo,empl)
        #print(empl.emp_performance)
        
    def create_sheet(self, depo, empl):
        curr_direc = os.getcwd()
        date = datetime.now().strftime('%Y_%m_%d_%H:%M:%S')
        filename = curr_direc + 'SDS_TN-' + date + '.xlsx'
        
        header_depo_day = ['Deň', 'Počet rozvozov', 'Hmotnosť rozvoz',
        'Počet zvozov', 'Hmotnosť zvoz', 'Σ (Rozvoz + zvoz)']
        header_depo_month = ['Mesiac', 'Počet rozvozov', 'Hmotnosť rozvoz',
        'Počet zvozov', 'Hmotnosť zvoz', 'Σ (Rozvoz + zvoz)']
        
        writer = pandas.ExcelWriter(filename,engine='xlsxwriter')
        workbook  = writer.book   
        header_fmt = workbook.add_format({'font_name': 'Arial', 'font_size': 10,
        'bold': True, 'bg_color': '#303030'})
        print(type(depo.month_list))
        depo_data = [depo.month_dataframe, depo.day_dataframe]
        row = 0
        for dataframe in depo_data:
            dataframe.to_excel(writer, 'Depo', startrow=row, startcol=0, index=False)
            if row == 0:
                worksheet = writer.sheets['Depo']
                worksheet.autofilter(row,0,row,5)
            row += len(dataframe.index) + 2 + 1
        
        #design
        worksheet.set_row(0,60)
        worksheet.set_column(0,0,10)
        worksheet.set_column(1,1,12)
        worksheet.set_column(2,2,13)
        worksheet.set_column(3,3,9)
        worksheet.set_column(4,4,12)
        worksheet.set_column(5,5,12)
        text_format = workbook.add_format({'text_wrap': True, "bold": True,
        'font_color': '#f1f1f1', "bg_color":"#303030", "valign":"vcenter", "align":"center"})
        
        for i in range (0, len(header_depo_month)):
            worksheet.write(0, i, header_depo_month[i], text_format)
        
            